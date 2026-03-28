import os
import re
from pathlib import Path
from langchain_community.document_loaders import Docx2txtLoader, TextLoader
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_core.documents import Document
from app.config import settings
from app.core.vectorstore import get_vectorstore

SUPPORTED_EXTENSIONS = {
    ".pdf", ".txt", ".md",
    ".docx", ".doc",
    ".xlsx", ".xls",
    ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff", ".tif", ".webp",
}


def _clean_text(text: str) -> str:
    """Clean text for DOCX/TXT — collapses whitespace."""
    text = re.sub(r'\b(?:[A-Za-z] ){3,}[A-Za-z]\b', lambda m: m.group(0).replace(' ', ''), text)
    text = re.sub(r"(\w)-\n(\w)", r"\1\2", text)
    text = re.sub(r"^\s*\d+\s*$", "", text, flags=re.MULTILINE)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r"[ \t]{2,}", " ", text)
    return text.strip()


def _annotate_financial_labels(text: str) -> str:
    """Convert ambiguous columnar financial layout into explicit LABEL: VALUE lines.

    PDFs like bank/credit card statements use multi-row column headers whose values
    appear on subsequent lines — models cannot reliably map them. This function
    detects these patterns and rewrites them as unambiguous key-value pairs BEFORE
    the text is chunked and embedded.
    """
    # TOTAL AMOUNT DUE — value appears within 3 lines after the label
    text = re.sub(
        r'TOTAL AMOUNT DUE\s*\n(?:[^\n]{0,80}\n){0,3}?\s*(?:_\s+)?((?:C|₹)\s*[\d,]+\.\d+)',
        r'TOTAL AMOUNT DUE: \1',
        text
    )
    # MINIMUM DUE — appears on a header line, value on next line
    text = re.sub(
        r'MINIMUM DUE\s+DUE DATE\s*\n\s*((?:C|₹)\s*[\d,]+\.\d+)\s+(\d{1,2}\s+\w+,?\s*\d{4})',
        r'MINIMUM DUE: \1\nDUE DATE: \2',
        text
    )
    # MINIMUM DUE without DUE DATE on same line
    text = re.sub(
        r'(?<!TOTAL )(?<!\w )MINIMUM DUE\s*\n\s*((?:C|₹)\s*[\d,]+\.\d+)',
        r'MINIMUM DUE: \1',
        text
    )
    # PREVIOUS STATEMENT DUES
    text = re.sub(
        r'PREVIOUS STATEMENT DUES?\s*\n\s*((?:C|₹)\s*[\d,]+\.\d+)',
        r'PREVIOUS STATEMENT DUES: \1',
        text
    )
    return text


def _clean_pdf_text(text: str) -> str:
    """Clean PDF text — moderate space collapse keeps column hints readable."""
    text = re.sub(r'\b(?:[A-Za-z] ){3,}[A-Za-z]\b', lambda m: m.group(0).replace(' ', ''), text)
    text = re.sub(r"(\w)-\n(\w)", r"\1\2", text)
    text = re.sub(r"^\s*\d+\s*$", "", text, flags=re.MULTILINE)
    # Remove standalone underscore placeholders (PDF visual separators)
    text = re.sub(r"^\s*_\s*$", "", text, flags=re.MULTILINE)
    text = re.sub(r"^\s*_\s+(?=\S)", "", text, flags=re.MULTILINE)
    text = re.sub(r"\n{3,}", "\n\n", text)
    # Collapse 4+ spaces to 2 — brings column values closer to headers
    text = re.sub(r" {4,}", "  ", text)
    # Apply financial label annotation after layout is normalised
    text = _annotate_financial_labels(text)
    return text.strip()


def _load_pdf(file_path: str) -> list[Document]:
    """PDF extraction using pdfplumber with layout=True.

    layout=True preserves horizontal column spacing so values stay aligned
    with their headers — critical for financial statements and tables.
    All pages are merged into one Document so chunks can span page boundaries.
    """
    import pdfplumber

    pages_content = []
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text(layout=True)
            if text and text.strip():
                pages_content.append(text.strip())

    if not pages_content:
        return []

    return [Document(
        page_content=_clean_pdf_text("\n\n".join(pages_content)),
        metadata={"source": file_path},
    )]


def _load_xls(file_path: str) -> list[Document]:
    """Extract legacy .xls sheets as structured pipe-delimited rows using xlrd."""
    import xlrd

    wb = xlrd.open_workbook(file_path)
    sections = []

    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        rows = []
        for rx in range(ws.nrows):
            cells = [str(ws.cell_value(rx, cx)).strip() for cx in range(ws.ncols)]
            if any(cells):
                rows.append(" | ".join(cells))
        if rows:
            sections.append(f"Sheet: {sheet_name}\n" + "\n".join(rows))

    if not sections:
        return []

    return [Document(
        page_content=_clean_text("\n\n".join(sections)),
        metadata={"source": file_path},
    )]


def _load_doc(file_path: str) -> list[Document]:
    """Extract legacy .doc files via antiword subprocess."""
    import subprocess

    try:
        result = subprocess.run(
            ["antiword", file_path],
            capture_output=True,
            text=True,
            timeout=30,
        )
        text = result.stdout.strip()
    except (FileNotFoundError, subprocess.TimeoutExpired):
        text = ""

    if not text:
        return []

    return [Document(
        page_content=_clean_text(text),
        metadata={"source": file_path},
    )]


def _fix_ocr_symbols(text: str) -> str:
    """Fix common Tesseract misreads of the Indian Rupee symbol ₹.

    Tesseract misreads ₹ as '2' or '£' depending on font rendering.
    We fix both cases when followed by Indian-format numbers.
    """
    # ₹ misread as '£' (most common with clear fonts)
    text = re.sub(r'£(\d{1,2},\d{2},\d{3}(?:\.\d+)?)', r'₹\1', text)   # lakh: £X,XX,XXX
    text = re.sub(r'£(\d{1,3},\d{3}(?:\.\d+)?)', r'₹\1', text)           # thousand: £X,XXX
    # ₹ misread as '2' (common with stylised/bold fonts)
    text = re.sub(r'(?<!\d)2(\d{1,2},\d{2},\d{3}(?:\.\d+)?)', r'₹\1', text)
    text = re.sub(r'(?<!\d)2(\d{1,2},\d{3}(?:\.\d+)?)', r'₹\1', text)
    return text


def _load_image(file_path: str) -> list[Document]:
    """Extract text from image files via Tesseract OCR."""
    import pytesseract
    from PIL import Image

    img = Image.open(file_path)
    text = pytesseract.image_to_string(img, config="--psm 6").strip()

    if not text:
        return []

    text = _fix_ocr_symbols(text)
    return [Document(
        page_content=_clean_text(text),
        metadata={"source": file_path},
    )]


def _load_excel(file_path: str) -> list[Document]:
    """Extract Excel sheets as structured pipe-delimited rows.

    Each sheet becomes a section. Headers are preserved so the model
    understands column-to-value relationships.
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            # Skip fully empty rows
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                rows.append(" | ".join(cells))
        if rows:
            sections.append(f"Sheet: {sheet_name}\n" + "\n".join(rows))

    if not sections:
        return []

    return [Document(
        page_content=_clean_text("\n\n".join(sections)),
        metadata={"source": file_path},
    )]


IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff", ".tif", ".webp"}


def load_document(file_path: str) -> list[Document]:
    ext = Path(file_path).suffix.lower()
    if ext == ".pdf":
        return _load_pdf(file_path)
    elif ext == ".xlsx":
        return _load_excel(file_path)
    elif ext == ".xls":
        return _load_xls(file_path)
    elif ext == ".doc":
        return _load_doc(file_path)
    elif ext in IMAGE_EXTENSIONS:
        return _load_image(file_path)
    elif ext == ".docx":
        loader = Docx2txtLoader(file_path)
    elif ext in {".txt", ".md"}:
        loader = TextLoader(file_path, encoding="utf-8")
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    docs = loader.load()
    for doc in docs:
        doc.page_content = _clean_text(doc.page_content)
    return docs


def chunk_documents(documents: list[Document]) -> list[Document]:
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=settings.chunk_size,
        chunk_overlap=settings.chunk_overlap,
        separators=["\n\n", "\n", ". ", " ", ""],
        length_function=len,
    )
    chunks = splitter.split_documents(documents)
    # Drop empty or near-empty chunks that produce bad embeddings
    return [c for c in chunks if len(c.page_content.strip()) > 50]


def ingest_file(file_path: str, doc_id: str, filename: str) -> dict:
    """Load → clean → chunk → embed → store in ChromaDB."""
    raw_docs = load_document(file_path)

    for doc in raw_docs:
        doc.metadata.update({"doc_id": doc_id, "filename": filename})

    chunks = chunk_documents(raw_docs)

    vs = get_vectorstore()
    chunk_ids = [f"{doc_id}_{i}" for i in range(len(chunks))]
    vs.add_documents(chunks, ids=chunk_ids)

    return {
        "doc_id": doc_id,
        "filename": filename,
        "total_pages": len(raw_docs),
        "total_chunks": len(chunks),
    }


def delete_document(doc_id: str) -> int:
    """Remove all chunks for a document from ChromaDB."""
    vs = get_vectorstore()
    collection = vs._collection
    results = collection.get(where={"doc_id": doc_id})
    ids_to_delete = results["ids"]
    if ids_to_delete:
        collection.delete(ids=ids_to_delete)
    return len(ids_to_delete)


def list_documents() -> list[dict]:
    """Return unique documents stored in ChromaDB."""
    vs = get_vectorstore()
    collection = vs._collection
    results = collection.get(include=["metadatas"])

    seen: dict[str, dict] = {}
    chunk_counts: dict[str, int] = {}

    for meta in results["metadatas"]:
        doc_id = meta.get("doc_id")
        if not doc_id:
            continue
        if doc_id not in seen:
            seen[doc_id] = {
                "doc_id": doc_id,
                "filename": meta.get("filename", "unknown"),
            }
            chunk_counts[doc_id] = 0
        chunk_counts[doc_id] += 1

    return [
        {**doc, "total_chunks": chunk_counts[doc["doc_id"]]}
        for doc in seen.values()
    ]
